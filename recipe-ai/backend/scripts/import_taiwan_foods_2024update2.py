"""
匯入「食品營養成分資料庫 2024 UPDATE2」Excel 到 PostgreSQL taiwan_foods 表。

Idempotent upsert by integration_code（Excel「整合編號」）：
- 既有 row 會更新到最新數值
- Excel 多出的（如 2,180 → 2,504）會新增

使用：
    docker cp .../食品營養成分資料庫2024UPDATE2.xlsx recipe-backend:/tmp/
    docker cp .../scripts recipe-backend:/app/scripts
    docker exec recipe-backend python /app/scripts/import_taiwan_foods_2024update2.py /tmp/食品營養成分資料庫2024UPDATE2.xlsx
"""
import asyncio
import os
import sys

import asyncpg
import openpyxl

# Excel 第 0 列是標題說明，第 1 列才是欄名，資料從第 2 列開始（zero-indexed）
HEADER_ROW = 2
DATA_START_ROW = 3

# Excel 中文欄名 → DB snake_case 欄名（按 Excel 欄位順序排列）
EXCEL_TO_DB = {
    "整合編號": "integration_code",
    "食品分類": "category",
    "樣品名稱": "food_name",
    "內容物描述": "sample_name",
    "俗名": "common_name",
    "廢棄率(%)": "waste_rate",
    "熱量(kcal)": "cal_per_100g",
    "修正熱量(kcal)": "modified_cal_per_100g",
    "水分(g)": "water_per_100g",
    "粗蛋白(g)": "protein_per_100g",
    "粗脂肪(g)": "fats_per_100g",
    "飽和脂肪(g)": "saturated_fat_per_100g",
    "灰分(g)": "ash_per_100g",
    "總碳水化合物(g)": "carbon_per_100g",
    "膳食纖維(g)": "dietary_fiber_per_100g",
    "糖質總量(g)": "total_sugar_per_100g",
    "葡萄糖(g)": "glucose_per_100g",
    "果糖(g)": "fructose_per_100g",
    "半乳糖(g)": "galactose_per_100g",
    "麥芽糖(g)": "maltose_per_100g",
    "蔗糖(g)": "sucrose_per_100g",
    "乳糖(g)": "lactose_per_100g",
    "鈉(mg)": "sodium_per_100g",
    "鉀(mg)": "potassium_per_100g",
    "鈣(mg)": "calcium_per_100g",
    "鎂(mg)": "magnesium_per_100g",
    "鐵(mg)": "iron_per_100g",
    "鋅(mg)": "zinc_per_100g",
    "磷(mg)": "phosphorus_per_100g",
    "銅(mg)": "copper_per_100g",
    "錳(mg)": "manganese_per_100g",
    "維生素A總量(IU)": "vitamin_a_total_iu",
    "視網醇當量(RE)(ug)": "retinol_equivalent_ug",
    "視網醇(ug)": "retinol_ug",
    "α-胡蘿蔔素(ug)": "alpha_carotene_ug",
    "β-胡蘿蔔素(ug)": "beta_carotene_ug",
    "維生素D總量(IU)": "vitamin_d_total_iu",
    "維生素D總量(ug)": "vitamin_d_total_ug",
    "維生素D2(ug)": "vitamin_d2_ug",
    "維生素D3(ug)": "vitamin_d3_ug",
    "維生素E總量(mg)": "vitamin_e_total_mg",
    "α-維生素E當量(α-TE)(mg)": "alpha_vitamin_e_te_mg",
    "α-生育酚(mg)": "alpha_tocopherol_mg",
    "β-生育酚(mg)": "beta_tocopherol_mg",
    "γ-生育酚(mg)": "gamma_tocopherol_mg",
    "δ-生育酚(mg)": "delta_tocopherol_mg",
    "維生素K1(ug)": "vitamin_k1_ug",
    "維生素K2 (MK-4)(ug)": "vitamin_k2_mk4_ug",
    "維生素K2 (MK-7)(ug)": "vitamin_k2_mk7_ug",
    "維生素B1(mg)": "vitamin_b1_mg",
    "維生素B2(mg)": "vitamin_b2_mg",
    "菸鹼素(mg)": "niacin_mg",
    "維生素B6(mg)": "vitamin_b6_mg",
    "維生素B12(ug)": "vitamin_b12_ug",
    "葉酸(ug)": "folate_ug",
    "維生素C(mg)": "vitamin_c_mg",
    "脂肪酸S總量(mg)": "saturated_fatty_acids_total_mg",
    "酪酸(4:0)(mg)": "butyric_acid_4_0_mg",
    "己酸(6:0)(mg)": "caproic_acid_6_0_mg",
    "辛酸(8:0)(mg)": "caprylic_acid_8_0_mg",
    "癸酸(10:0)(mg)": "capric_acid_10_0_mg",
    "月桂酸(12:0)(mg)": "lauric_acid_12_0_mg",
    "十三酸(13:0)(mg)": "tridecanoic_acid_13_0_mg",
    "肉豆蔻酸(14:0)(mg)": "myristic_acid_14_0_mg",
    "十五酸(15:0)(mg)": "pentadecanoic_acid_15_0_mg",
    "棕櫚酸(16:0)(mg)": "palmitic_acid_16_0_mg",
    "十七酸(17:0)(mg)": "heptadecanoic_acid_17_0_mg",
    "硬脂酸(18:0)(mg)": "stearic_acid_18_0_mg",
    "十九酸(19:0)(mg)": "nonadecanoic_acid_19_0_mg",
    "花生酸(20:0)(mg)": "arachidic_acid_20_0_mg",
    "山酸(22:0)(mg)": "behenic_acid_22_0_mg",
    "廿四酸(24:0)(mg)": "lignoceric_acid_24_0_mg",
    "脂肪酸M總量(mg)": "monounsaturated_fatty_acids_total_mg",
    "肉豆蔻烯酸(14:1)(mg)": "myristoleic_acid_14_1_mg",
    "棕櫚烯酸(16:1)(mg)": "palmitoleic_acid_16_1_mg",
    "油酸(18:1)(mg)": "oleic_acid_18_1_mg",
    "鱈烯酸(20:1)(mg)": "gadoleic_acid_20_1_mg",
    "芥子酸(22:1)(mg)": "erucic_acid_22_1_mg",
    "脂肪酸P總量(mg)": "polyunsaturated_fatty_acids_total_mg",
    "亞麻油酸(18:2)(mg)": "linoleic_acid_18_2_mg",
    "次亞麻油酸(18:3)(mg)": "linolenic_acid_18_3_mg",
    "十八碳四烯酸(18:4)(mg)": "octadecatetraenoic_acid_18_4_mg",
    "花生油酸(20:4)(mg)": "arachidonic_acid_20_4_mg",
    "廿碳五烯酸(20:5)(mg)": "epa_20_5_mg",
    "廿二碳五烯酸(22:5)(mg)": "dpa_22_5_mg",
    "廿二碳六烯酸(22:6)(mg)": "dha_22_6_mg",
    "其他脂肪酸(mg)": "other_fatty_acids_mg",
    "P/M/S": "p_m_s_ratio",
    "反式脂肪(mg)": "trans_fat_mg",
    "水解胺基酸總量(mg)": "total_amino_acids_mg",
    "天門冬胺酸(Asp)(mg)": "aspartic_acid_mg",
    "酥胺酸(Thr)(mg)": "threonine_mg",
    "絲胺酸(Ser)(mg)": "serine_mg",
    "麩胺酸(Glu)(mg)": "glutamic_acid_mg",
    "脯胺酸(Pro)(mg)": "proline_mg",
    "甘胺酸(Gly)(mg)": "glycine_mg",
    "丙胺酸(Ala)(mg)": "alanine_mg",
    "胱胺酸(Cys)(mg)": "cystine_mg",
    "纈胺酸(Val)(mg)": "valine_mg",
    "甲硫胺酸(Met)(mg)": "methionine_mg",
    "異白胺酸(Ile)(mg)": "isoleucine_mg",
    "白胺酸(Leu)(mg)": "leucine_mg",
    "酪胺酸(Tyr)(mg)": "tyrosine_mg",
    "苯丙胺酸(Phe)(mg)": "phenylalanine_mg",
    "離胺酸(Lys)(mg)": "lysine_mg",
    "組胺酸(His)(mg)": "histidine_mg",
    "精胺酸(Arg)(mg)": "arginine_mg",
    "色胺酸(Trp)(mg)": "tryptophan_mg",
    "膽固醇(mg)": "cholesterol_mg",
    "酒精含量(g)": "alcohol_content_g",
}

TEXT_COLS = {"integration_code", "category", "food_name", "sample_name", "common_name", "p_m_s_ratio"}


def to_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "—", "N/A", "n/a", "ND", "nd", "tr", "Tr", "TR"):
        return None
    if s.startswith("<"):
        # 「<0.1」之類視為缺值（也可改成 0，但保守處理用 NULL）
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


async def main(xlsx_path: str):
    print(f"[1/4] 讀取 Excel: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # 讀第 2 列為 header（zero-indexed row 1，openpyxl 是 1-indexed 的 row 2）
    header_row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    headers = [str(h).strip() if h is not None else None for h in header_row]

    # 建 col_index → db_col_name 映射；遇到不認得的欄名印出警告
    col_to_db = {}
    for idx, h in enumerate(headers):
        if not h:
            continue
        if h in EXCEL_TO_DB:
            col_to_db[idx] = EXCEL_TO_DB[h]
        else:
            print(f"  ⚠ 略過未知欄位 [{idx}] {h!r}")
    print(f"  → 對應 {len(col_to_db)} 個欄位（預期 110）")

    # 建立 row dict list
    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if all(v is None for v in row):
            continue
        rec = {}
        for idx, db_col in col_to_db.items():
            v = row[idx] if idx < len(row) else None
            rec[db_col] = to_text(v) if db_col in TEXT_COLS else to_number(v)
        # 必要：integration_code 必須存在才能 upsert
        if not rec.get("integration_code") or not rec.get("food_name"):
            continue
        rows.append(rec)
    wb.close()
    print(f"[2/4] 解析 {len(rows)} 筆有效資料（必須有 integration_code 與 food_name）")

    # 建 SQL（依 EXCEL_TO_DB 的 value 集合作為欄位順序，確保 upsert 完整覆蓋）
    db_cols = list(dict.fromkeys(EXCEL_TO_DB.values()))  # preserve order, dedupe
    placeholders = ",".join(f"${i+1}" for i in range(len(db_cols)))
    update_clause = ",".join(f"{c}=EXCLUDED.{c}" for c in db_cols if c != "integration_code")
    sql = (
        f"INSERT INTO taiwan_foods ({','.join(db_cols)}) "
        f"VALUES ({placeholders}) "
        f"ON CONFLICT (integration_code) DO UPDATE SET {update_clause}, updated_at=NOW() "
        f"RETURNING (xmax = 0) AS inserted"
    )

    # 連線：優先 DATABASE_URL，其次採用 backend 容器的 BACKEND_* env（與 app_config.py 對齊）
    db_url = os.environ.get("DATABASE_URL")
    if db_url:
        print(f"[3/4] 連線 DB: {db_url.split('@')[-1]}")
        conn = await asyncpg.connect(db_url)
    else:
        host = os.environ.get("BACKEND_DB_HOST", "localhost")
        port = int(os.environ.get("BACKEND_DB_PORT", "5432"))
        database = os.environ.get("BACKEND_DB_NAME", "recipe_ai")
        user = os.environ.get("BACKEND_DB_USER", "postgres")
        password = os.environ.get("BACKEND_DB_PASSWORD", "")
        print(f"[3/4] 連線 DB: {user}@{host}:{port}/{database}")
        conn = await asyncpg.connect(
            host=host, port=port, database=database, user=user, password=password
        )

    inserted = 0
    updated = 0
    try:
        # 確認 integration_code UNIQUE constraint 存在；migration 002 已宣告 UNIQUE，但保險再 ensure
        await conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_taiwan_foods_integration_code "
            "ON taiwan_foods(integration_code)"
        )

        async with conn.transaction():
            for rec in rows:
                params = [rec.get(c) for c in db_cols]
                row = await conn.fetchrow(sql, *params)
                if row and row["inserted"]:
                    inserted += 1
                else:
                    updated += 1

        total_now = await conn.fetchval("SELECT COUNT(*) FROM taiwan_foods")
        print(f"[4/4] INSERTED={inserted}  UPDATED={updated}  TOTAL_NOW={total_now}")
    finally:
        await conn.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python import_taiwan_foods_2024update2.py <path/to/xlsx>", file=sys.stderr)
        sys.exit(1)
    asyncio.run(main(sys.argv[1]))
